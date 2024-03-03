# Standard imports
import os
import glob
import threading
from typing import Tuple
import icecream as ic

# Data Science imports
import numpy as np
import pandas as pd
from scipy.signal import argrelextrema

# visualisation imports
import mplfinance as mpf
import matplotlib.pyplot as plt

# excel imports
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# PricePatternPro imports
from headAndShoulders_inverse import headAndShoulders_inverse
from intraday_data import Intra_Histdata_Request
from intraday_data import TradingApp

from Context import Context
from doubleBottom import doubleBottom
from doubleTop import doubleTop
from headAndShoulders import headAndShoulders
from tripleTop import tripleTop


def main():

    context = Context(tripleTop())

    # Initialize variables
    minima, maxima= {}, {}
    pattern_data = pd.DataFrame()

    strategy_data_global = {}
    trade_summary_global = pd.DataFrame()

    # Read historical data files
    histdata = read_Daily_historical_files()

# =============================================================================
#     few_pairs = {}
#     count = 0
#     for key, value in histdata.items():
#         if count < 3:
#             few_pairs[key] = value
#             count += 1
#         else:
#             break
# =============================================================================

    for key , value in histdata.items():
        print("-"*60)
        print(f"ticker = {key}")
        print("-"*15)
        if key == "BTDR":
            x=10
        #identify all the low and highs swing points on all charts
        minima, maxima = get_min_max(value)

        pattern_data = context.pattern_detection(minima, maxima, value)

        # if pattern is not confirmed then skip ticker, start next loop
        if len(pattern_data) == 0:
            continue

        plot_pattern(value, pattern_data, key, context)

        #Prepare backtesting
        strategy_data = pre_backtester(value, pattern_data )
        #Start backtesting
        trade_summary = backtester(strategy_data, key)
        #Post backtesting
        post_backtester(strategy_data, trade_summary, context._bias)

        # Prepare Intraday backtest
        intra_histdata = intraday_pre_backtester(trade_summary, key)
        # Intraday Backtest
        intraday_backtest(intra_histdata, trade_summary)

        # Trading Cost and Slippage Modeling
        trade_summary['total_charges'] = compute_Trading_Fees(intra_histdata, trade_summary, key)

        # Store Trade data in a global variable
        trade_summary_global = pd.concat([trade_summary_global, trade_summary])
        # Store strategy data in a global variable
        strategy_data_global[key] = strategy_data

        plot_trades(value, trade_summary, key, context._name)

    # Portfolio & Strategy Performance Anylysis
    trade_summary_global.reset_index(drop=True, inplace = True)

    #log_trades(trade_summary_global)

    # Analyse strategy performance
    strategy_performance = startegy_level_analysis(trade_summary_global)

    #Analyze portfolio performance
    portfolio_performance = portfolio_level_analysis(strategy_data_global, trade_summary_global, context._name)

    # Combined performance
    performance = pd.concat([strategy_performance, portfolio_performance])

    x = 0

###############################################################################
############################## GENERAL FUNCTIONS ##############################
###############################################################################

def get_min_max(df :pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    argrel_window = 5

    #use the argrelextrema to compute the local minima and maxima points
    local_min = argrelextrema(df.iloc[:-argrel_window]['Low'].values,
                              np.less, order=argrel_window)[0]

    local_max = argrelextrema(df.iloc[:-argrel_window]['High'].values,
                              np.greater, order=argrel_window)[0]

    #store the minima and maxima values in a dataframe
    return  df.iloc[local_min].Low,  df.iloc[local_max].High

def plot_pattern(data: pd.DataFrame, pattern_data: pd.DataFrame, ticker: str, context: Context):
    points = context._datapoints
    bias = context._bias
    name = context._name

    for x in range(0, len(pattern_data)):
        data_to_plot =  pattern_data.iloc[x]

        # Capture the pattern details for the most recent pattern
        dt = pd.DataFrame()
        dt['dates'] = data_to_plot.iloc[:points]
        dt['price'] = data_to_plot.iloc[points:2*points].values

        # Initialise the start and end dates for OHLCV data
        start_index = data.index.get_loc(dt.at[dt.index[0],'dates']) - 3
        if start_index < 0:
            start_index = 0
        last_index = data.index.get_loc(pattern_data.at[x,'confirmation_date']) + 3
        if last_index > len(data):
            last_index = len(data)

        # Fetch the OHLCV data for the period
        data_to_plot = data.iloc[start_index:last_index]

        # Add pattern details to data_to_plot
        pd.set_option('mode.chained_assignment', None)
        data_to_plot.loc[dt['dates'],'pattern'] = dt['price'].values
        data_to_plot.loc[pattern_data.at[x,'confirmation_date'],'trade'] = data.loc[pattern_data.at[x,'confirmation_date'], 'Close']
        pd.set_option('mode.chained_assignment', 'warn')

        # Define scatter plot to mark the sequential points forming the pattern
        dt_plot = [mpf.make_addplot(data_to_plot['pattern'], type='scatter', marker='o', markersize=400, panel=0, color="blue", alpha=0.7),
                   mpf.make_addplot(data_to_plot['trade'], type='scatter', marker='o', markersize=400, panel=0, color="red", alpha=0.7),]

        if bias == "Long":
            start_price = data.iloc[start_index:data.index.get_loc(dt.at[dt.index[0],'dates'])].max()["High"]
            end_price = data.iloc[data.index.get_loc(dt.at[dt.index[-1],'dates']):last_index].max()["High"]
        elif bias == "Short":
            start_price = data.iloc[start_index:data.index.get_loc(dt.at[dt.index[0],'dates'])].min()["Low"]
            end_price = data.iloc[data.index.get_loc(dt.at[dt.index[-1],'dates']):last_index].min()["Low"]

        # Define subplot of a line that will connect the 5 sequential points
        pattern_line = [(data_to_plot.index[0], start_price)] \
                     + list(dt.itertuples(index=False, name=None)) \
                     + [(data_to_plot.index[-1], end_price)]

        title = str(x) +' - ' + ticker + ' - ' + name

        # Plot the candlestick chart along with pattern-related subplots
        mpf.plot(data_to_plot, type='candle', style='classic', addplot=dt_plot,
                 title=title, figsize=(15, 7), alines=pattern_line,
                 hlines=dict(hlines=[pattern_data.at[x,'target'],
                                     pattern_data.at[x,'stoploss']],
                             colors=['g', 'r'], linestyle='-.'))

def plot_trades(data: pd.DataFrame, trade_data: pd.DataFrame, ticker: str, pattern_name = str):

    data.loc[trade_data['entry_date'].dt.date,'trade_entry'] = trade_data['entry_price'].values
    data.loc[trade_data['exit_date'].dt.date,'trade_exit'] = trade_data['exit_price'].values

    # Define scatter plot to mark the entry and exit trades
    dt_plot = [mpf.make_addplot(data['trade_entry'], type='scatter', marker='>', markersize=300, panel=0, color="green", alpha=0.7),
               mpf.make_addplot(data['trade_exit'], type='scatter', marker='<', markersize=300, panel=0, color="red", alpha=0.7),]

    title = ticker + ' - ' + pattern_name

    # Plot the candlestick chart along with pattern-related subplots
    mpf.plot(data, type='candle', style='classic', addplot=dt_plot,
             title=title, figsize=(15, 7))

###############################################################################
######################## READ HISTORICAL DATA FUNCTIONS #######################
###############################################################################

def read_Daily_historical_files() ->dict :
    # change directory
    os.chdir("C:\\X\\Workspaces\\ALGO\\Data\\Historical\\Daily")

    # Collect all filenames in current directory
    filenames = glob.glob("*_Daily_Bars.csv")

    histdata = { } #empty dictionary NO KEYS NO VALUES

    for filename in filenames:
        # open, read and store content in a dataframe
        with open("C:\\X\\Workspaces\\ALGO\\Data\\Historical\\Daily\\" + filename, "r") as file:
            df = pd.read_csv(file,index_col=0)
            df.index = pd.to_datetime(df.index, format='%Y%m%d')

        # round each price column
        for i in range (0,4):
            df.iloc[:,i] = round(df.iloc[:,i],2)

        # store data in a dict of dataframes
        ticker = filename.split("_")[0]
        histdata[ticker] = df

    return histdata

def read_Intra_historical_files(filenames:list) -> dict:

    # change directory
    os.chdir("C:\\X\\Workspaces\\ALGO\\Data\\Historical\\Intraday")

    #filenames = glob.glob("*_Intraday_Bars.csv")

    histdata = { }

    for filename in filenames:
        # open, read and store content in a dataframe
        with open("C:\\X\\Workspaces\\ALGO\\Data\\Historical\\Intraday\\" + filename, "r") as file:
            df = pd.read_csv(file,index_col=0)
            df.index = pd.to_datetime(df.index)

        # round each price column
        for i in range (0,4):
            df.iloc[:,i] = round(df.iloc[:,i],2)

        # store data in a dict of dataframes
        ticker = filename.split("_")[0]
        histdata[ticker] = df

    return histdata

###############################################################################
########################## DAILY BACKTEST FUNCTIONS ###########################
###############################################################################

def pre_backtester(data: pd.DataFrame, pattern_data: pd.DataFrame) -> pd.DataFrame:

    pattern_data['confirmation_date'] = pd.to_datetime(pattern_data['confirmation_date'], format='%Y-%m-%d')

    #if confirmation date is duplicated for 2 patterns remove it
    pattern_data = pattern_data.drop_duplicates(subset=['confirmation_date'], keep='first')

    pattern_data = pattern_data.reset_index(drop = True)

    #start merging
    strategy_data = pd.merge(data, pattern_data, how ='left',
                                left_on = 'Date', right_on ='confirmation_date')

    #index was lost in merging, retrieve it
    strategy_data.index  = data.index

    # replace na with 0 in signal column
    strategy_data.signal.fillna(0, inplace=True)

    return strategy_data

def backtester(data: pd.DataFrame, ticker: str) -> pd.DataFrame:
    # ------------------------------- Initial Settings ------------------------
    # Create dataframes for round trips, storing trades, and mtm
    trade_summary = pd.DataFrame()
    trade = pd.DataFrame()

    # Initilise current position, number of trades, cumulative pnl, stop-loss
    # to 0 and take ptofit to 100000
    current_position = 0
    trade_num = 0
    cum_pnl = 0
    stop_loss = 0
    take_profit = 100000

    # Set exit and entry flag to False
    exit_flag = False
    entry_flag = False

    # ------------------------------- Backtest Over Historical Data -----------


    for i in data.index:
        # ------------------- Positions Check ----------------------------
        # No positions
        if (current_position == 0) & (data.loc[i,'signal'] != 0):
            current_position = data.loc[i,'signal']
            entry_flag = True

        # Short positions
        elif current_position == -1:
            if (data.loc[i,'signal'] != 0):
                data.loc[i,'signal'] = 0

            if data.loc[i,'Close'] > stop_loss:
                exit_type = 'SL'
                exit_flag = True

            elif data.loc[i,'Close'] < take_profit:
                exit_type = 'TP'
                exit_flag = True

            elif data.index[-2] == i:
                if data.loc[i,'Close'] < trade.entry_price[0]:
                    exit_type = 'LP'
                elif data.loc[i,'Close'] > trade.entry_price[0]:
                    exit_type = 'LL'
                exit_flag = True

        # Long positions
        elif current_position == 1:
            if (data.loc[i,'signal'] != 0):
                data.loc[i,'signal'] = 0

            if data.loc[i,'Close'] < stop_loss:
                exit_type = 'SL'
                exit_flag = True

            elif data.loc[i,'Close'] > take_profit:
                exit_type = 'TP'
                exit_flag = True

            elif data.index[-2] == i:
                if data.loc[i,'Close'] > trade.entry_price[0]:
                    exit_type = 'LP'
                elif data.loc[i,'Close'] < trade.entry_price[0]:
                    exit_type = 'LL'
                exit_flag = True

        # ------------------------------- Entry Position Update ---------------
        if entry_flag:

            #Populate the trades dataframe
            trade = pd.DataFrame(index = [0])
            trade['Symbol'] = ticker
            trade['entry_date'] = i
            trade['entry_price'] = round(data.loc[i,'Close'],2)
            trade['position'] = current_position

            stop_loss = data.loc[i,'stoploss']

            take_profit = data.loc[i,'target']

            trade_num += 1

            # Print trade details
            print(f"\033[36mTrade No: {trade_num}\033[0m | Entry Date: {i} | Entry Price : {trade.entry_price[0]}] | Position: {current_position}  | Symbol: {ticker}")

            # Set entry flag to False
            entry_flag = False

            continue

        # ------------------------------- Exit Position Update ----------------
        if exit_flag:

            #Populate the trades dataframe
            trade['exit_date'] = i
            trade['exit_type'] = exit_type
            trade['exit_price'] = round(data.loc[i,'Close'],2)

            # Calculate pnl for the trade
            trade_pnl = current_position *  round( trade.exit_price[0] - \
                                                   trade.entry_price[0],2)

            # Calculate cumulative pnl
            cum_pnl += trade_pnl
            cum_pnl = round(cum_pnl,2)
            trade['PnL'] = trade_pnl


            trade['stoploss'] = stop_loss
            trade['target'] = take_profit

            # Add the trade logs to round trip details
            trade_summary = pd.concat([trade_summary, trade])

            # Print trade details
            print(f"Exit Type: {exit_type} | Exit Date: {i} | Exit Price: {trade.exit_price[0]}   | PnL: {trade_pnl} | Cum PnL: {cum_pnl}")
            print("-"*30)

            # Update current position to 0
            current_position = 0

            # Set the exit flag to False
            exit_flag = False

            continue

    trade_summary.reset_index(drop=True, inplace=True)

    return trade_summary

def post_backtester(strategy_data: pd.DataFrame, trade_summary:pd.DataFrame, bias: str) -> pd.DataFrame:

    if bias == "Long":
        # exit is a sell orderd
        strategy_data.loc[ trade_summary['exit_date'] , 'signal'] = -1
    elif bias == "Short":
        # exit is a buy order
        strategy_data.loc[ trade_summary['exit_date'] , 'signal'] = 1

###############################################################################
########################## INTRADAY BACKTEST FUNCTIONS ########################
###############################################################################

def Request_Intraday_Historical_Data(entry_n_exit_dates: list):
    event = threading.Event()

    app_obj = TradingApp(event)

    req = Intra_Histdata_Request(event, app_obj)
    req.request(entry_n_exit_dates)

def intraday_pre_backtester (trade_details: pd.DataFrame, ticker: str)->dict:
    entry_n_exit_dates = []
    file_list = []

    result_list = []
    for idx, row in trade_details.iterrows():

        symbol = row['Symbol']
        #entry_date
        entry_date = row['entry_date'].strftime('%Y%m%d')
        entry_n_exit_dates.append( [ symbol, entry_date ] )
        file_list.append(symbol +'-'+ entry_date +'-entry_Intraday_Bars.csv')

        #exit_date
        exit_date = row['exit_date'].strftime('%Y%m%d')
        entry_n_exit_dates.append( [ symbol, exit_date  ] )
        file_list.append(symbol +'-'+ exit_date + '-exit_Intraday_Bars.csv')

    # Send Requests to IBAPI
    #Request_Intraday_Historical_Data(entry_n_exit_dates)

    # Read intraday historical data files
    intra_histdata = read_Intra_historical_files(file_list)

    # perform sanity check on data
    #is_data_clean = data_sanity_check(intra_histdata)

    return intra_histdata

def intraday_backtest(data: dict, trades: pd.DataFrame) -> pd.DataFrame:

    #trades = trades.set_index('Symbol')
    # ------------------------------- Backtest Over Historical Data -------------------------------------
    for key , df in  data.items():
        ticker, date, action = key.split('-')
        exit_found_flag, entry_found_flag = False, False

        for i in df.index:
            # ------------------- Entry position check ----------------------------
            if action == 'entry':
                # Since my entry was based on the closing price of my daily candle
                # Then entry date is set to the last candle in intraday data
                # and entry price is not optimized

                idx = trades.loc[trades['entry_date'] == date].index[0]

                # Short positions
                if entry_found_flag == False and trades.at[idx, 'position'] == -1:
                    trades.at[idx, 'entry_date'] = data[key].index[-2]
                    entry_found_flag = True
                    break

                # Long position
                elif entry_found_flag == False and trades.at[idx, 'position'] == 1:
                    trades.at[idx, 'entry_date'] = data[key].index[-2]
                    entry_found_flag = True
                    break


            # ------------------- Exit position check ----------------------------
            elif action == 'exit':

                idx = trades.loc[trades['exit_date'] == date].index[0]

                # Short position
                if trades.at[idx, 'position'] == -1:

                    # EXIT == STOPLOSS
                    if exit_found_flag == False and data[key].at[i, 'Close'] >= trades.at[idx, 'stoploss']:
                        trades.at[idx, 'exit_date'] = i
                        trades.at[idx,'exit_price'] = round(data[key].loc[i,'Close'],2)
                        exit_found_flag = True
                        break

                    # EXIT  == TAKE PROFIT
                    elif exit_found_flag == False and data[key].at[i,'Close'] <= trades.at[idx, 'target']:
                        trades.at[idx, 'exit_date'] = i
                        trades.at[idx,'exit_price'] = round(data[key].loc[i,'Close'],2)
                        exit_found_flag = True
                        break

                    # EXIT == LOCK LOSS OR LOCK PROFIT
                    elif exit_found_flag == False and (trades.at[idx, 'exit_type'] == 'LP' or trades.at[idx, 'exit_type'] == 'LL'):
                        trades.at[idx, 'exit_date'] = data[key].index[-2]
                        trades.at[idx,'exit_price'] = round(data[key].loc[i,'Close'],2)
                        exit_found_flag = True
                        break

                # Long position
                elif trades.at[idx, 'position'] == 1:

                    # EXIT == STOPLOSS
                    if exit_found_flag == False and data[key].at[i, 'Close'] <= trades.at[idx, 'stoploss']:
                        trades.at[idx, 'exit_date'] = i
                        trades.at[idx,'exit_price'] = round(data[key].loc[i,'Close'],2)
                        exit_found_flag = True
                        break

                    # EXIT  == TAKE PROFIT
                    elif exit_found_flag == False and data[key].at[i,'Close'] >= trades.at[idx, 'target']:
                        trades.at[idx, 'exit_date'] = i
                        trades.at[idx,'exit_price'] = round(data[key].loc[i,'Close'],2)
                        exit_found_flag = True
                        break

                    # EXIT == LOCK LOSS OR LOCK PROFIT
                    elif exit_found_flag == False and (trades.at[idx, 'exit_type'] == 'LP' or trades.at[idx, 'exit_type'] == 'LL'):
                        trades.at[idx, 'exit_date'] = data[key].index[-2]
                        trades.at[idx,'exit_price'] = round(data[key].loc[i,'Close'],2)
                        exit_found_flag = True
                        break

###############################################################################
########################## TRADING COSTS FUNCTIONS  ###########################
###############################################################################

def compute_Trading_Fees(intra_histdata: dict, trades: pd.DataFrame, ticker: str) -> pd.DataFrame:

    # Step 1: Define other costs
    transaction_cost = 0.0001
    broker_cost = 0.03/100
    tax_cost = 0.02/100
    slippage_cost = slippage_Modelling(intra_histdata, trades)# Model for Slippage

    # Step 2: Combine other costs into a DataFrame
    total_charges = pd.DataFrame({
                            'transaction_cost': [transaction_cost] * len(slippage_cost),
                            'broker_cost': [broker_cost] * len(slippage_cost),
                            'tax_cost': [tax_cost] * len(slippage_cost)  },
                        index=slippage_cost.index)

    # Step 3: Combine slippage_cost DataFrame with other costs
    total_charges = pd.concat([slippage_cost, total_charges], axis=1)

    # Step 4: Sum each row
    total_charges['total_charges'] = total_charges.sum(axis=1)

    return total_charges['total_charges']

def slippage_Modelling(intra_histdata: dict, trades:pd.DataFrame) -> pd.DataFrame:
    dates = []
    last_5_candles = {}
    slippage_df = pd.DataFrame()

    # store entry and exit dates conscutivley in a list
    for idx, row in trades.iterrows():
        dates.append(row['entry_date'])
        dates.append(row['exit_date'])

    ctr = 0

    for key, df in intra_histdata.items():

        action = key.split('-')[2]

        if action == 'entry':
            #get index integer location for the entry date
            index = df.index.get_loc( dates[ctr] ) +1

            #slice last 5 candles
            df = df.reset_index()
            last_5_candles[key] = df.iloc[index-5:index]
            df = df.set_index('Date')

            # For a buy order, Worst execution price (high price) - Last traded price (close price)
            # For a sell order, Last traded price (close price) - Worst execution price (low price)

            #------------------ Slippage for entry orders -------------------------

            # Short Position (enty = sell order)
            if trades.at[ trades.index[int(ctr/2)] , 'position'] == -1:

                last_5_candles[key]['slippage_entry_order'] = \
                    (last_5_candles[key]['Close'] - last_5_candles[key]['Low']) \
                        / last_5_candles[key]['Close']

            # Long Position (entry = buy order)
            elif trades.at[ trades.index[int(ctr/2)] , 'position'] == 1:

                last_5_candles[key]['slippage_entry_order'] = \
                    (last_5_candles[key]['High'] - last_5_candles[key]['Close']) \
                        / last_5_candles[key]['Close']

            # Calculate average of slippage over last 5 candles
            slippage_df.at[trades.index[int(ctr/2)]  ,'slippage_entry_order'] = \
                last_5_candles[key]['slippage_entry_order'].mean()

        elif action == 'exit':
            #get index integer location for the exit date
            index = df.index.get_loc( dates[ctr] ) +1
            if (index < 5):
                index = 5

            # slice last 5 candles
            df = df.reset_index()
            last_5_candles[key] = df.iloc[index-5:index]
            df = df.set_index('Date')

            #------------------ Slippage for exit orders --------------------------

            # Short Position (exit = buy order)
            if trades.at[trades.index[int(ctr/2)], 'position'] == -1:
                last_5_candles[key]['slippage_exit_order'] = \
                    (last_5_candles[key]['High'] - last_5_candles[key]['Close']) \
                        / last_5_candles[key]['Close']

            # Long Position (exit = sell order)
            elif trades.at[trades.index[int(ctr/2)], 'position'] == 1:
                last_5_candles[key]['slippage_exit_order'] = \
                    (last_5_candles[key]['Close'] - last_5_candles[key]['Low']) \
                        / last_5_candles[key]['Close']

            # Calculate average of slippage over last 5 candles
            slippage_df.at[trades.index[int(ctr/2)],'slippage_exit_order'] = \
                last_5_candles[key]['slippage_exit_order'].mean()

        ctr += 1

    return slippage_df

###############################################################################
########################## TRADING LOGS FUNCTIONS   ###########################
###############################################################################

def log_trades(trades_summary: pd.DataFrame):

    file_path = "C:\\X\\Workspaces\\ALGO\Data\\trades\\Strategy_Performance.xlsx"

    wb = load_workbook(file_path, read_only = True, data_only=True)
    # Select worksheet
    ws = wb['Double Top']

    # Convert DataFrame to rows
    rows = list(dataframe_to_rows(trades_summary, index=False, header=False))

    # Update cells in the worksheet
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, value in enumerate(row, start=2):
            ws.cell(row=r_idx, column=c_idx, value=value)



    # Specify the row containing column names
    wb = load_workbook(file_path)
    excel_data = pd.read_excel(file_path, header = 1)

    # Find the index of the column 'Price Out' (assuming it exists in your DataFrame)
    column_index = excel_data.columns.get_loc('Price Out')

    # Keep only the columns before 'Price Out'
    excel_data = excel_data.iloc[:, :column_index + 1]

    excel_column_names = {'entry_date'  : 'Date',
                          'Symbol'      : 'Ticker',
                          'position'    : 'Position',
                          'entry_price' : 'Price In',
                          'exit_price'  : 'Price Out'  }

    # Create a Pandas Excel writer using XlsxWriter as the engine
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Write the DataFrame to a specific sheet named 'Trades' starting from cell B2
        trades_summary.to_excel(writer, sheet_name='Double Top', startrow=3, startcol=2, index=False)

        # Adjust column locations based on excel_column_names
        for df_col, excel_col in excel_column_names.items():
            col_index = excel_data.columns.get_loc(excel_col)
            df_col_index = trades_summary.columns.get_loc(df_col)

            for i, val in enumerate(trades_summary[df_col]):
                excel_data.iloc[i, col_index] = val


    print("Data saved to Excel file:", file_path)

###############################################################################
#################### PERFORMANCE ANALYSIS FUNCTIONS ###########################
###############################################################################

def startegy_level_analysis(trades: pd.DataFrame) -> pd.DataFrame:

    # Create dataframe to store trade analytics
    analytics = pd.DataFrame(index=['Strategy'])

    #for key, value in trades.items:

    # Calculate total Pnl
    analytics['Total Pnl'] = trades.PnL.sum()

    analytics['Cost'] = round(trades['total_charges'].sum(), 4 )

    # Number of total trades
    analytics['total_trades'] = len(trades)

    # Profitable trades
    analytics['Number of winners'] = len(trades.loc[trades.PnL>0])
    # Loosing trades
    analytics['Number of losers'] = len(trades.loc[trades.PnL<0])

    # Win Percentage
    analytics['Win (%)'] = 100 * analytics['Number of winners']/analytics.total_trades

    # Loss Percentage
    analytics['Loss (%)'] = 100 * analytics['Number of losers']/analytics.total_trades

    # Per trade profit/loss of winning trades
    analytics['per_trade_PnL_winners'] = round(trades.loc[trades.PnL>0].PnL.mean(),2)

    # Per trade profit/loss of losing trades
    analytics['per_trade_PnL_losers'] = round(np.abs(trades.loc[trades.PnL<0].PnL.mean()),2)

    # Calculate holding period for each trade
    holding_period = trades['exit_date'] - trades['entry_date']

    # Calculate mean of holding period
    analytics['Average holding time'] = holding_period.mean()

    # Calculate profit factor
    analytics['Profit Factor'] = (analytics['Win (%)' ]/100*analytics['per_trade_PnL_winners']) / \
                                 (analytics['Loss (%)']/100*analytics['per_trade_PnL_losers'] )

    return analytics.T

def portfolio_level_analysis(strategy_data: dict, trades: pd.DataFrame, pattern_name: str):

    performance_metrics = pd.DataFrame(index=['Strategy'])
    cumulative_return = pd.DataFrame()

    # get equity curve for each individual stock
    for stock, df in strategy_data.items():
        fees = round(trades[ trades['Symbol'] == stock ].total_charges.sum(),2)

        cumulative_return[stock] = compute_EquityCurve(df, fees)

    # aggregate all equity curves into a single curve
    cumulative_return['Strategy'] = cumulative_return.sum(axis=1) - (len(cumulative_return.columns) - 1)

    # drop first row
    cumulative_return = cumulative_return.dropna()

    # -------------------Plot Equity Curve ------------------------------------
    # Set the title and axis labels
    plt.plot(cumulative_return['Strategy'], color='purple')
    plt.title(f' {pattern_name} - Portfolio Equity Curve', fontsize = 14)
    plt.ylabel('Cumulative Returns', fontsize = 12)
    plt.xlabel('year', fontsize = 12)
    plt.show()

    # ------ Compund Annual Growth Rate (CAGR) --------------------------------
    performance_metrics['CAGR'] = compute_CAGR( cumulative_return['Strategy'] )

    # ---------------- Annualized Volatility ----------------------------------
    performance_metrics['Annualised Volatility'] = compute_annualized_volatility( cumulative_return['Strategy'] )

    # ------------------- Sharpe Ratio ----------------------------------------
    performance_metrics['Sharpe Ratio'] = compute_SharpeRatio( cumulative_return['Strategy'] )

    # ------------------- Maximum Drawdown ------------------------------------
    performance_metrics['Maximum Drawdown'] = compute_DrawDown( cumulative_return['Strategy'] , pattern_name)

    return performance_metrics.T

###############################################################################
########################## COMPUTATION FUNCTIONS ##############################
###############################################################################

def compute_EquityCurve(data, total_charges):
    """

    """
    # calculate the percentage change for each close price. Shift signal to avoid lookahead bias
    data['stock_return'] = data['Close'].pct_change() * data['signal'].shift(1)

    # Calculate the trading cost when you square off the position
    trading_cost = (total_charges * np.abs(data['signal'] - data['signal'].shift(1)))

    # Calculate net strategy returns
    data['strategy_returns_minus_cost'] = data['stock_return'] - trading_cost

    # 1 is the initial capital added on top of the stock return
    data['cumulative_returns'] = (1 +  data['strategy_returns_minus_cost'] ).cumprod()

    return data['cumulative_returns']

def compute_CAGR(strategy_equity_curve):
    """
    It is the measure of a strategy's annual growth rate over time
    """
    # Calculate beginning and ending portfolio values
    beginning_portfolio_value = strategy_equity_curve.iloc[0]
    ending_portfolio_value = strategy_equity_curve.iloc[-1]

    # Calculate number of years (assuming daily returns)
    num_years = len(strategy_equity_curve) / 252  # Assuming 252 trading days per year

    # Calculate CAGR
    cagr = ( (ending_portfolio_value / beginning_portfolio_value) ** (1 / num_years) - 1 ) * 100

    return round(cagr, 2)

def compute_SharpeRatio(strategy_equity_curve):
    """
    It is the ratio of the returns earned in excess of the risk-free rate per unit of risk taken
    """
    # Calculate number of years (assuming daily returns)
    n = len(strategy_equity_curve)/252 # Assuming 252 trading days per year

    # Set a risk-free rate
    risk_free_rate = 0.02/(252*n)

    # Calculate Sharpe ratio
    Sharpe_Ratio = np.sqrt(252*n) * (strategy_equity_curve.mean() - (risk_free_rate)) / strategy_equity_curve.std()

    return round(Sharpe_Ratio, 2)

def compute_DrawDown(strategy_equity_curve, pattern_name = str):
    """
    It signifies the maximum loss from a peak to a trough of a strategy and is expressed in percentage terms.
    """
    data = pd.DataFrame()

    # Compute the cumulative maximum
    data['Peak'] = strategy_equity_curve.cummax()

    # Compute the Drawdown
    data['Drawdown'] = ( ( strategy_equity_curve - data['Peak'] )/ data['Peak'] ) * 100

    # Compute the maximum drawdown
    data['Maximum Drawdown'] = data['Drawdown'].min()

    # -------------------Plot maximum drawdown---------------------
    plt.plot(data['Drawdown'], alpha = 0.5, color='red' )
    plt.fill_between(data['Drawdown'].index, data['Drawdown'].values, color='red')
    plt.title(f'{pattern_name} - Portfolio Drawdown', fontsize=14)
    plt.ylabel('Drawdown(%)', fontsize=12)
    plt.xlabel('Year', fontsize=12)
    plt.show()

    return round(data['Maximum Drawdown'].min(),2)

def compute_annualized_volatility(strategy_equity_curve):
    """
    Volatility is the rate at which the strategy returns increases or decreases over a year.
    """
    annualized_volatiliy = strategy_equity_curve.std()*np.sqrt(252) * 100

    return round(annualized_volatiliy,2)


if __name__ == "__main__":
    main()